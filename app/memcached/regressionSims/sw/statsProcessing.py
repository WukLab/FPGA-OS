#-------------------------------------------------------------------------------
# Name:        emmaOutputParser
# Purpose:
#
# Author:      kimonk
#
# Created:     15/11/2012
#-------------------------------------------------------------------------------

import time
import math
import sys
import string
from pylab import *
from openpyxl import Workbook
from openpyxl.chart import LineChart, Serie, Reference

statsItem = []
statsList = []

if (len(sys.argv) !=2):
    print("Error! Incorrect number of arguments passed.")
    sys.exit(2)
else:
    iFileName = sys.argv[2]

counter = 0
inputFile = open(iFileName, 'r')
for iLine in inputFile:

    line = iLine.split()
    if (line[1] == "2"):
        statsItem.append(line[0])
        if(counter == 13):
            counter = 0
            statsList.append(statsItem)
            del statsItem[:]
        else:
            counter += 1
    else:
        continue
#Finished reading in of the stats items and pushed them into the statsList. Need
#to plot them now on graphs
#Rearrdange the elements
plotList = []
for i in range(0, 13):
    for item in statsList:
        plotList.append(item[i])

inputFile.close()

################################################################################
################################################################################
################################################################################
    for
            fig10 = figure()
            maxXaxis = max(maxReadLatency, maxWriteLatency)
            xAxis = [i for i in range (0, maxXaxis)]
            plot(xAxis, writeLatencyDistr[0:maxXaxis], 'bo')
            plot(xAxis, readLatencyDistr[0:maxXaxis], 'ro')
            xlabel("Latency (ns)")
            ylabel("Occurance")
            title("Read/Write Latency distribution for packets of size "+ str(packetsize))
            # show()
            if(saveImage == 1): savefig("latencyDistro_it" + str(counter), bbox_inches=0)
            latencyDistroFigures.append(fig10)
            del fig10



        # Parse the emma schedule file
        resultsFile.write("----- Emma Schedule Analysis - Channel " + str(k) + "----\n")
        resultsFile.write("# of ACT: " + str(sum(activates)) + "\n")
        resultsFile.write("# of PRE: " + str(sum(precharges)) + "\n")
        resultsFile.write("# of WR: " + str(sum(writes)) + "\n")
        resultsFile.write("# of WRA: " + str(sum(wrAs)) + "\n")
        resultsFile.write("# of RD: " + str(sum(reads)) + "\n")
        resultsFile.write("# of RDA: " + str(sum(rdAs)) + "\n")
        cmds = sum(reads) + sum(writes) + sum (rdAs) + sum(wrAs)
        resultsFile.write("# of commands executed: " + str(cmds) + "\n")
        usedBusCycles = (cmds * (burstLength/2))
        resultsFile.write("# of used bus cycles: " + str(usedBusCycles) + "\n")
        memEfficiency.append(((cmds * (burstLength/2)) * 100)/noOfCycles)
        resultsFile.write("Mem. Efficiency: " + str(memEfficiency[k]) + "%" + "\n")
        throughput.append(((hbmFreq*busWidth*2)*memEfficiency[k])/(100*1000))
        resultsFile.write("Throughput: " + str(throughput[k]) + " Gbps" + "\n")
        # Then output the same stuff per bank
        resultsFile.write("-------------- Statistics per bank ----------------\n")
        for j in range(0, noOfBanks):
            resultsFile.write("Bank " + str(j) + ":\n")
            resultsFile.write("# of ACT: " + str(activates[j]) + "\n")
            resultsFile.write("# of PRE: " + str(precharges[j]) + "\n")
            resultsFile.write("# of WR: " + str(writes[j]) + "\n")
            resultsFile.write("# of WRA: " + str(wrAs[j]) + "\n")
            resultsFile.write("# of RD: " + str(reads[j]) + "\n")
            resultsFile.write("# of RDA: " + str(rdAs[j]) + "\n")
        # Output latency related information
        resultsFile.write("-------------- Latency information ----------------\n")
        if(latencyCounter != 0):
            resultsFile.write("Average latency per mem. access: " + str(latencyAccumulate / latencyCounter) + "\n")
            avgLatency = latencyAccumulate / latencyCounter
        else:
            resultsFile.write("Average latency for all commands: 0")
        if (writeLatencyCounter != 0):
           resultsFile.write("Average latency for writes: " + str(writeLatencyAccumulate / writeLatencyCounter) + "\n")
           avgWriteLatency = writeLatencyAccumulate / writeLatencyCounter
        else:
           resultsFile.write("Average latency for writes: 0")
        if (readLatencyCounter != 0):
           resultsFile.write("Average latency for reads: "+ str(readLatencyAccumulate / readLatencyCounter) + "\n")
           avgReadLatency = readLatencyAccumulate / readLatencyCounter
        else:
           resultsFile.write("Average latency for reads: 0")
        #Set all counter variables to 0
        tgReadCounter, tgWriteCounter, tgWriteBytes, tgReadBytes, tgActualWriteBytes, tgActualReadBytes = 0, 0, 0, 0, 0, 0
        latencyAccumulate = 0
        latencyCounter = 0
        writeLatencyAccumulate = 0
        writeLatencyCounter = 0
        readLatencyAccumulate = 0
        readLatencyCounter = 0
        activates = [0 for x in activates]  # Set all the counter variables to 0 before the next iteration
        precharges = [0 for x in precharges]
        writes = [0 for x in writes]
        reads = [0 for x in reads]
        wrAs = [0 for x in wrAs]
        rdAs = [0 for x in rdAs]
    #append the lists
    itThroughput.append(throughput)
    throughput = []
    itMemEfficiency.append(memEfficiency)
    memEfficiency = []
    itLatencyAVG.append(latencyAVG)
    latencyAVG = []
    itWriteLatencyAVG.append(writeLatencyAVG)
    writeLatencyAVG = []
    itReadLatencyAVG.append(readLatencyAVG)
    readLatencyAVG = []
    resultsFile.close()
        # write batch mem.efficiency file
    for k in range(0, len(itMemEfficiency[counter])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        datapoints.write(str(itMemEfficiency[counter][k]) + " ")
        latencypoints.write(str(itLatencyAVG[counter][k]) + " ")
    #print(str(counter))
    counter += 1
cscript.close()
datapoints.write("\n")
datapoints.close()
# Create the plots
# Plot throughput
time.sleep(10)
fig1 = figure()
if (bwLatency == 2):
    plot(flowNumber, packetsDropped, 'bo-')
    xlabel("No. of flows")
    title("Dropped packets for various no. of flows")
elif (bwLatency == 1):
    plot(simThroughput, packetsDropped, 'bo-')
    xlabel("Input Bandwdith in Gbps")
    title("Dropped packets for different input rates")
elif (bwLatency == 0):
    plot(packetSizes, packetsDropped, 'bo-')
    xlabel("Packet Size in Bytes")
    title("Dropped packets for different mem. access sizes")
ylabel("% of packets dropped")
if (saveImage == 1): savefig("packetDrop", bbox_inches=0)
# Plot mem. throughput
tempThroughput = [] # Temporary list to rearrange the values read from the files
for k in range(0, len(itThroughput[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
    for i in range(0, len(itThroughput)):
        tempThroughput.append(itThroughput[i][k])
    fig2 = figure()
    if(bwLatency == 2):
        plot(flowNumber, tempThroughput, 'bo-')
        xlabel("No. of flows")
        title("Mem. Throughput for various no. of flows - Channel " + str(k+1))
    elif(bwLatency == 1):
        plot(simThroughput, tempThroughput, 'bo-')
        xlabel("Input Bandwdith in Gbps")
        title("Mem. Throughput for different input rates - Channel " + str(k+1))
    elif (bwLatency == 0):
        plot(packetSizes, tempThroughput, 'bo-')
        xlabel("Packet Size in Bytes")
        title("Mem. Throughput for different mem. access sizes - Channel " + str(k+1))
    ylabel("Throughput in Gbps")
    if (saveImage == 1): savefig("memThroughput_ch" + str(k+1), bbox_inches=0)
    tempThroughput = []
#del tempThroughput
# Plot mem. efficiency
tempMemEfficiency = [] # Temporary list to rearrange the values read from the files
for k in range(0, len(itMemEfficiency[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
    for i in range(0, len(itMemEfficiency)):
        tempMemEfficiency.append(itMemEfficiency[i][k])
    fig3 = figure()
    if(bwLatency == 2):
        plot(flowNumber, tempMemEfficiency, 'bo-')
        xlabel("No. of flows")
        title("Mem. bus utilization for various no. of flows - Channel " + str(k+1))
    elif(bwLatency == 1):
        plot(simThroughput, tempMemEfficiency, 'bo-')
        xlabel("Input Bandwdith in Gbps")
        title("Mem. bus utilization for different input rates")
        tempSimThroughput = []
        for i in range(0, len(simThroughput)):
            tempSimThroughput.append(int(simThroughput[i]))
        xMax = max(tempSimThroughput)
        del tempSimThroughput
        xlim([0, int(xMax)])
    elif (bwLatency == 0):
        plot(packetSizes, tempMemEfficiency, 'bo-')
        xlabel("Packet Size in Bytes")
        title("Mem. bus utilization for different mem. access sizes - Channel " + str(k+1))
    ylabel("Mem. Bus Utilization (in %)")
    ylim([0, 100])
    if (saveImage == 1): savefig("memEfficiency_ch" + str(k+1), bbox_inches=0)
    tempMemEfficiency = []
#del tempMemEfficiency
# Plot total avg. latency
tempLatencyAVG = [] # Temporary list to rearrange the values read from the files
avgLatencyArr = []
for k in range(0, len(itLatencyAVG[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
    for i in range(0, len(itLatencyAVG)):
        tempLatencyAVG.append(itLatencyAVG[i][k])
        avgLatencyArr.append(np.mean(itLatencyAVG))
    fig3 = figure()
    if(bwLatency == 2):
        plot(flowNumber, tempLatencyAVG, 'bo-')
        xlabel("No. of flows")
        title("Latency for various no. of flows")
    elif(bwLatency == 1):
        plot(simThroughput, tempLatencyAVG, 'bo-')
        xlabel("Input Bandwdith in Gbps")
        title("Bandwidth Latency graph for different input rates - Channel " + str(k+1))
        xlim([0, int(xMax)])
        ylim([0, 100])
    elif (bwLatency == 0):
        plot(packetSizes, tempLatencyAVG, 'bo-')
        plot(packetSizes, avgLatencyArr, 'r-')
        xlabel("Packet Size in Bytes")
        title("Latency graph  for different mem. access sizes - Channel " + str(k+1))
        ylim([0, max(tempLatencyAVG)])
    ylabel("Average mem. access latency in ns")
    if (saveImage == 1): savefig("bwLatency_ch" + str(k+1), bbox_inches=0)
    tempLatencyAVG = []
#del tempLatencyAVG
# Plot total avg. latency for reads
tempReadLatencyAVG = [] # Temporary list to rearrange the values read from the files
avgReadLatencyArr = []
for k in range(0, len(itReadLatencyAVG[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
    for i in range(0, len(itReadLatencyAVG)):
        tempReadLatencyAVG.append(itReadLatencyAVG[i][k])
        avgReadLatencyArr.append(np.mean(itReadLatencyAVG))
    fig4 = figure()
    if(bwLatency == 2):
        plot(flowNumber, tempReadLatencyAVG, 'bo-')
        xlabel("No. of flows")
        title("Read latency for various no. of flows")
    elif(bwLatency == 1):
        plot(simThroughput, tempReadLatencyAVG, 'bo-')
        xlabel("Input Bandwdith in Gbps")
        title("Bandwidth Latency graph over different input rates - Channel " + str(k+1))
        xlim([0, int(xMax)])
        ylim([0, 100])
    elif (bwLatency == 0):
        plot(packetSizes, tempReadLatencyAVG, 'bo-')
        plot(packetSizes, avgReadLatencyArr, 'r-')
        xlabel("Packet Size in Bytes")
        title("Read latency graph  for different mem. access sizes - Channel " + str(k+1))
        ylim([0, max(tempReadLatencyAVG)])
    ylabel("Average read access latency in ns")
    if (saveImage == 1): savefig("bwLatencyRD_ch" + str(k+1), bbox_inches=0)
    tempReadLatencyAVG = []
#del tempReadLatencyAVG
# Plot total avg. latency for writes
tempWriteLatencyAVG = [] # Temporary list to rearrange the values read from the files
avgWriteLatencyArr = []
for k in range(0, len(itWriteLatencyAVG[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
    for i in range(0, len(itWriteLatencyAVG)):
        tempWriteLatencyAVG.append(itWriteLatencyAVG[i][k])
        avgWriteLatencyArr.append(np.mean(itWriteLatencyAVG))
    fig5 = figure()
    if(bwLatency == 2):
        plot(flowNumber, tempWriteLatencyAVG, 'bo-')
        xlabel("No. of flows")
        title("Write latency for various no. of flows")
    elif(bwLatency == 1):
        plot(simThroughput, tempWriteLatencyAVG, 'bo-')
        xlabel("Input Bandwdith in Gbps")
        title("Bandwidth Latency graph over different input rates - Channel " + str(k+1))
        xlim([0, int(xMax)])
        ylim([0, 100])
    elif (bwLatency == 0):
        plot(packetSizes, tempWriteLatencyAVG, 'bo-')
        plot(packetSizes, avgWriteLatencyArr, 'r-')
        xlabel("Packet Size in Bytes")
        title("Bandwidth Latency graph  for different mem. access sizes - Channel " + str(k+1))
        ylim([0, max(tempWriteLatencyAVG)])
    ylabel("Average write access latency in ns")
    if (saveImage == 1): savefig("bwLatencyWR_ch" + str(k+1), bbox_inches=0)
    if (showGraphs == 1): show()
    tempWriteLatencyAVG = []
#del tempWriteLatencyAVG
#draw()
# Conditional Excel Output
if(createExcel == 1):
    wb = Workbook() # Creates a workbook in memory
    ws1 = wb.get_active_sheet() # Get the active worksheet
    ws1.title = "Throughput & Mem. Efficiency"
    for k in range(0, len(itThroughput[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "A" + str(3+(k*2))
        ws1.cell(cellName).value = "Channel " + str(k)
        ws1.cell(cellName).style.font.bold = True
        cellName = "B" + str(3+(k*2))
        ws1.cell(cellName).value = "Throughput"
        ws1.cell(cellName).style.font.bold = True
        cellName = "B" + str(4+(k*2))
        ws1.cell(cellName).value = "Mem. Efficiency"
        ws1.cell(cellName).style.font.bold = True
    if (bwLatency == 2):
        for c,n in zip(ws1.range("C2:N2")[0], flowNumber):
            c.value = n
            c.style.font.bold = True
    elif (bwLatency == 1):
        for c,n in zip(ws1.range("C2:N2")[0], simThroughput):
            c.value = n
            c.style.font.bold = True
    elif (bwLatency == 0):
        for c,n in zip(ws1.range("C2:N2")[0], packetSizes):
            c.value = n
            c.style.font.bold = True


    for k in range(0, len(itThroughput[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "C" + str(3+(k*2)) + ":N" + str(3+(k*2))
        for i in range(0, len(itThroughput)):
            tempThroughput.append(itThroughput[i][k])
        for c,n in zip(ws1.range(cellName)[0], tempThroughput):
        #for c,n in zip(ws1.range("C3:N3")[0], tempThroughput):
            c.value = n
        tempThroughput = []

    for k in range(0, len(itMemEfficiency[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "C" + str(4+(k*2)) + ":N" + str(4+(k*2))
        for i in range(0, len(itMemEfficiency)):
            tempMemEfficiency.append(itMemEfficiency[i][k])
        #for c,n in zip(ws1.range("C4:N4")[0], tempMemEfficiency):
        for c,n in zip(ws1.range(cellName)[0], tempMemEfficiency):
            c.value = n
        tempMemEfficiency = []

    ws2 = wb.create_sheet() # Creates a new datasheet for the bandwidth latency graph
    for k in range(0, len(itThroughput[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "A" + str(3+(k*3))
        ws2.cell(cellName).value = "Channel " + str(k)
        ws2.cell(cellName).style.font.bold = True
        cellName = "B" + str(3+(k*3))
        ws2.title = "Bandwidth-Latency Diagrams"
        ws2.cell(cellName).value = "Avg. Latency"
        ws2.cell(cellName).style.font.bold = True
        ws2.cell(cellName).value = "Avg. Read Latency"
        ws2.cell(cellName).style.font.bold = True
        ws2.cell(cellName).value = "Avg. Write Latency"
        ws2.cell(cellName).style.font.bold = True

    if (bwLatency == 2):
        for c,n in zip(ws2.range("C2:N2")[0], flowNumber):
            c.value = n
            c.style.font.bold = True
    elif (bwLatency == 1):
        for c,n in zip(ws2.range("C2:N2")[0], simThroughput):
            c.value = n
            c.style.font.bold = True
    elif (bwLatency == 0):
        for c,n in zip(ws2.range("C2:N2")[0], packetSizes):
            c.value = n
            c.style.font.bold = True

    for k in range(0, len(itLatencyAVG[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "C" + str(3+(k*3)) + ":N" + str(3+(k*3))
        for i in range(0, len(itLatencyAVG)):
            tempLatencyAVG.append(itLatencyAVG[i][k])
        for c,n in zip(ws2.range(cellName)[0], tempLatencyAVG):
            c.value = n
        tempLatencyAVG = []

    for k in range(0, len(itReadLatencyAVG[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "C" + str(4+(k*3)) + ":N" + str(4+(k*3))
        for i in range(0, len(itReadLatencyAVG)):
            tempReadLatencyAVG.append(itReadLatencyAVG[i][k])
        for c,n in zip(ws2.range(cellName)[0], tempReadLatencyAVG):
            c.value = n
        tempReadLatencyAVG = []

    for k in range(0, len(itWriteLatencyAVG[0])): # go through all the simulated channels for the rest of the files (there is only one tg for all channels)
        cellName = "C" + str(5+(k*3)) + ":N" + str(5+(k*3))
        for i in range(0, len(itWriteLatencyAVG)):
            tempWriteLatencyAVG.append(itWriteLatencyAVG[i][k])
        for c,n in zip(ws2.range(cellName)[0], tempWriteLatencyAVG):
            c.value = n
        tempWriteLatencyAVG = []

    ws3 = wb.create_sheet()
    ws3.title = "Packet Drop Rate"

    ws3.cell('A4').value = "Packet Drop Rate (in %)"
    ws3.cell('A4').style.font.bold = True
    if (bwLatency == 2):
        ws3.cell('A3').value = "No. of Flows"
        for c,n in zip(ws2.range("B2:M2")[0], flowNumber):
            c.value = n
            c.style.font.bold = True
    elif (bwLatency == 1):
        ws3.cell('A3').value = "Input Bandwidth"
        for c,n in zip(ws2.range("B2:M2")[0], simThroughput):
            c.value = n
            c.style.font.bold = True
    elif (bwLatency == 0):
        ws3.cell('A3').value = "Packet Size"
        for c,n in zip(ws2.range("B2:M2")[0], packetSizes):
            c.value = n
            c.style.font.bold = True
    ws3.cell('A3').style.font.bold = True
    for c, n in zip(ws3.range("B4:M4")[0], packetsDropped):
        c.value = n

    wb.save("emmaResults.xlsx")
    # Throughput and mem. efficiency charts
    throughputChart = LineChart()
    throughputChart.show_legend = False
    if(bwLatency == 1): throughputChart.x_axis.title = "No. of Flows"
    elif(bwLatency == 1): throughputChart.x_axis.title = "Input Throughput (in Gbps)"
    elif(bwLatency == 0): throughputChart.x_axis.title = "Packet Size (in Bytes)"
    throughputChart.y_axis.title = "Mem. Throughput (in Gbps)"
    throughputChart.add_serie(Serie(Reference(ws1, (2, 2), (2, 2 + counter)), labels=(Reference(ws1, (1, 1), (1, counter)))))
    ws1.add_chart(throughputChart)
    memEfficienyChart = LineChart()
    memEfficienyChart.y_axis.min = 0
    memEfficienyChart.y_axis.max = 100
    memEfficienyChart.add_serie(Serie(Reference(ws1, (3, 2), (3, 2 + counter)), labels=(Reference(ws1, (1, 1), (1, counter)))))
    memEfficienyChart.show_legend = False
    memEfficienyChart.y_axis.title = "Mem. Efficiency (in % of clock cycles used)"
    if(bwLatency == 2): memEfficienyChart.x_axis_title = "No. of Flows"
    elif(bwLatency == 1): memEfficienyChart.x_axis.title = "Input Throughput (in Gbps)"
    elif(bwLatency == 0): memEfficienyChart.x_axis.title = "Packet Size (in Bytes)"
    memEfficienyChart.x_axis.title = "Input Throughput (in Gbps)"
    ws1.add_chart(memEfficienyChart)
    #Latency Charts
    latencyChart = LineChart() # chart for the average Latency per access for all mem. accesses
    latencyChart.add_serie(Serie(Reference(ws2, (2, 2), (2, 2 + counter)), labels=(Reference(ws2, (1, 1), (1, counter)))))
    latencyChart.show_legend = False
    latencyChart.y_axis.title = "Avg. Latency (in ns)"
    latencyChart.y_axis.min = 0
    latencyChart.y_axis.max = 100
    latencyChart.y_axis.unit = 10
    if(bwLatency == 2): latencyChart.x_axis_title = "No. of Flows"
    elif(bwLatency == 1): latencyChart.x_axis.title = "Input Throughput (in Gbps)"
    elif(bwLatency == 0): latencyChart.x_axis.title = "Packet Size (in Bytes)"
    ws2.add_chart(latencyChart)
    readLatencyChart = LineChart() # chart for the average Latency per access for reads only
    readLatencyChart.add_serie(Serie(Reference(ws2, (2, 2), (2, 2 + counter)), labels=(Reference(ws2, (1, 1), (1, counter)))))
    readLatencyChart.show_legend = False
    readLatencyChart.y_axis.title = "Avg. Read Latency (in ns)"
    readLatencyChart.y_axis.min = 0
    readLatencyChart.y_axis.max = 100
    readLatencyChart.y_axis.unit = 10
    if(bwLatency == 2): readLatencyChart.x_axis_title = "No. of Flows"
    elif(bwLatency == 1): readLatencyChart.x_axis.title = "Input Throughput (in Gbps)"
    elif(bwLatency == 0): readLatencyChart.x_axis.title = "Packet Size (in Bytes)"
    ws2.add_chart(readLatencyChart)
    writeLatencyChart = LineChart() # chart for the average Latency per access for writes only
    writeLatencyChart.add_serie(Serie(Reference(ws2, (2, 2), (2, 2 + counter)), labels=(Reference(ws2, (1, 1), (1, counter)))))
    writeLatencyChart.show_legend = False
    writeLatencyChart.y_axis.title = "Avg. Write Latency (in ns)"
    writeLatencyChart.y_axis.min = 0
    writeLatencyChart.y_axis.max = 100
    writeLatencyChart.y_axis.unit = 10
    if(bwLatency == 2): writeLatencyChart.x_axis_title = "No. of Flows"
    elif(bwLatency == 1): writeLatencyChart.x_axis.title = "Input Throughput (in Gbps)"
    elif(bwLatency == 0): writeLatencyChart.x_axis.title = "Packet Size (in Bytes)"
    ws2.add_chart(writeLatencyChart)
    # Packet Drop Rate Charts
    dropChart = LineChart()
    dropChart.add_serie(Serie(Reference(ws2, (2, 2), (2, 2 + counter)), labels=(Reference(ws2, (1, 1), (1, counter)))))
    dropChart.show_legend = False
    dropChart.y_axis.title = "Packet drop rate (in %)"
    dropChart.y_axis.min = 0
    dropChart.y_axis.max = 100
    dropChart.y_axis.unit = 10
    if(bwLatency == 2): dropChart.x_axis_title = "No. of Flows"
    elif(bwLatency == 1): dropChart.x_axis.title = "Input Throughput (in Gbps)"
    elif(bwLatency == 0): dropChart.x_axis.title = "Packet Size (in Bytes)"
    wb.save("emmaResults.xlsx")